"""Iceland energy data — generation by source from the Energy Authority XLSX archive.

Usage:
    uv run python scripts/energy.py list
    uv run python scripts/energy.py fetch
"""
from __future__ import annotations

import argparse
from pathlib import Path

import httpx
import polars as pl
from openpyxl import load_workbook

GENERATION_URL = (
    "https://vefskrar.orkustofnun.is/Talnaefni/"
    "OS-2025-1-throun-raforkuframleidslu-a-islandi-1969-2024.xlsx"
)
RAW_DIR = Path("data/raw/energy")
OUT = Path("data/processed/energy_generation.parquet")


def parse_generation(path: Path) -> pl.DataFrame:
    """Extract the first year-by-generation-source table from the workbook.

    The Authority periodically changes publication names, but the worksheet's
    table is self-describing: locate a row headed ``Ár``/``Year`` then read
    contiguous numeric year rows beneath it instead of hard-coding row numbers.
    """
    ws = load_workbook(path, data_only=True, read_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    header_i = next(
        i for i, row in enumerate(rows)
        if any(str(v).strip().lower() in {"ár", "year"} for v in row if v is not None)
    )
    header = list(rows[header_i])
    year_col = next(i for i, v in enumerate(header) if str(v).strip().lower() in {"ár", "year"})
    units = list(rows[header_i + 2])
    # The first table also contains an MWh duplicate total and percentage growth
    # calculations. Keep only physical generation sources plus the GWh total.
    source_cols = {
        i
        for i, label in enumerate(header)
        if i != year_col
        and (
            (i < len(units) and units[i] == "GWh" and str(label).strip() != "Samtals [GWh]")
            or str(label).strip() == "Samtals [GWh]"
        )
    }
    records: list[dict[str, object]] = []
    for row in rows[header_i + 1 :]:
        year = row[year_col] if year_col < len(row) else None
        if not isinstance(year, (int, float)) or not 1900 <= int(year) <= 2100:
            if records:
                break
            continue
        for col, label in enumerate(header):
            value = row[col] if col < len(row) else None
            if col not in source_cols or value is None or not isinstance(value, (int, float)):
                continue
            name = "Samtals" if str(label).strip() == "Samtals [GWh]" else str(label).strip()
            records.append({"year": int(year), "series": name, "gwh": float(value)})
    if not records:
        raise ValueError("generation table had no numeric rows")
    return pl.DataFrame(records).sort(["year", "series"])


def cmd_list(_: argparse.Namespace) -> None:
    print("generation\t1969–2024\tXLSX\t" + GENERATION_URL)


def cmd_fetch(_: argparse.Namespace) -> None:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    raw = RAW_DIR / "generation_1969_2024.xlsx"
    with httpx.Client(timeout=60, follow_redirects=True) as client:
        response = client.get(GENERATION_URL)
        response.raise_for_status()
    raw.write_bytes(response.content)
    df = parse_generation(raw)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    df.write_parquet(OUT)
    print(f"  {len(df):,} generation observations, {df['year'].min()}–{df['year'].max()}")
    print(f"  Wrote {raw} and {OUT}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(required=True)
    sub.add_parser("list").set_defaults(func=cmd_list)
    sub.add_parser("fetch").set_defaults(func=cmd_fetch)
    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
