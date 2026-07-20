"""Health probe — Energy Authority's working generation workbook."""
from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from scripts.energy import GENERATION_URL


def test_generation_workbook_has_year_series(http):
    r = http.get(GENERATION_URL)
    assert r.status_code == 200, f"{r.request.url} -> {r.status_code}"
    assert r.content[:2] == b"PK", "generation download is not an XLSX/ZIP file"
    ws = load_workbook(BytesIO(r.content), read_only=True, data_only=True).active
    values = [cell.value for row in ws.iter_rows(max_row=100, max_col=12) for cell in row]
    assert any(str(v).strip().lower() in {"ár", "year"} for v in values if v is not None)
    assert any(isinstance(v, (int, float)) and 2000 <= v <= 2100 for v in values)
